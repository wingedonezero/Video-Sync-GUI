#!/usr/bin/env python3
"""
VSG Rewrite Tracker Generator

Generates a comprehensive Excel spreadsheet for tracking the VSG Python application rewrite.
Examines all Python files in the 'Reference Only original' folder and extracts:
- Classes and their methods
- Standalone functions
- Key features and capabilities

Creates multiple sheets organized by functional area with:
- Progress tracking formulas
- Conditional formatting
- Status/Priority dropdowns
"""

import ast
import os
import re
from pathlib import Path
from typing import List, Dict, Any, Tuple
from dataclasses import dataclass, field
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule, DataBarRule
from openpyxl.worksheet.datavalidation import DataValidation


@dataclass
class CodeItem:
    """Represents a function, class, or feature to track."""
    file_path: str
    name: str
    item_type: str  # 'class', 'function', 'method', 'feature'
    description: str
    parent_class: str = ""
    features: List[str] = field(default_factory=list)
    line_number: int = 0


def extract_docstring(node) -> str:
    """Extract docstring from an AST node."""
    if isinstance(node.body, list) and node.body:
        first = node.body[0]
        if isinstance(first, ast.Expr) and isinstance(first.value, ast.Constant):
            doc = first.value.value
            if isinstance(doc, str):
                # Clean up and truncate
                doc = doc.strip().split('\n')[0][:200]
                return doc
    return ""


def extract_features_from_code(source: str, func_name: str) -> List[str]:
    """Extract notable features from function source code."""
    features = []
    source_lower = source.lower()

    # Pattern matching for common features
    feature_patterns = [
        (r'single.*(file|source|mode)', 'Handles single file mode'),
        (r'batch|multiple|loop.*for', 'Supports batch processing'),
        (r'async|await|asyncio', 'Async/await support'),
        (r'thread|threading|concurrent', 'Multi-threaded'),
        (r'cache|memoize', 'Caching support'),
        (r'validate|validation', 'Includes validation'),
        (r'retry|backoff', 'Retry logic'),
        (r'fallback', 'Fallback handling'),
        (r'error|exception|try:', 'Error handling'),
        (r'log|logging|_log', 'Logging'),
        (r'ffmpeg|ffprobe', 'FFmpeg integration'),
        (r'mkvmerge|mkvextract', 'MKVToolNix integration'),
        (r'subprocess|popen', 'Subprocess execution'),
        (r'json\.load|json\.dump', 'JSON handling'),
        (r'path|pathlib', 'Path handling'),
        (r'signal|emit', 'Qt signals'),
        (r'widget|dialog|window', 'UI component'),
    ]

    for pattern, feature in feature_patterns:
        if re.search(pattern, source_lower):
            if feature not in features:
                features.append(feature)

    return features[:5]  # Limit to 5 features


def analyze_python_file(file_path: Path, base_path: Path) -> List[CodeItem]:
    """Analyze a Python file and extract all classes, functions, and methods."""
    items = []
    relative_path = str(file_path.relative_to(base_path))

    try:
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
            source = f.read()

        tree = ast.parse(source)
    except (SyntaxError, UnicodeDecodeError) as e:
        # Still add an entry for the file
        items.append(CodeItem(
            file_path=relative_path,
            name=file_path.stem,
            item_type='file',
            description=f'Error parsing: {str(e)[:50]}',
        ))
        return items

    # Extract classes and their methods
    for node in ast.walk(tree):
        if isinstance(node, ast.ClassDef):
            class_doc = extract_docstring(node)
            items.append(CodeItem(
                file_path=relative_path,
                name=node.name,
                item_type='class',
                description=class_doc or f'Class definition',
                line_number=node.lineno,
            ))

            # Extract methods
            for item in node.body:
                if isinstance(item, ast.FunctionDef) or isinstance(item, ast.AsyncFunctionDef):
                    method_doc = extract_docstring(item)
                    # Get method source for feature extraction
                    try:
                        method_source = ast.get_source_segment(source, item) or ""
                        features = extract_features_from_code(method_source, item.name)
                    except:
                        features = []

                    items.append(CodeItem(
                        file_path=relative_path,
                        name=item.name,
                        item_type='method',
                        description=method_doc or f'Method of {node.name}',
                        parent_class=node.name,
                        features=features,
                        line_number=item.lineno,
                    ))

    # Extract module-level functions
    for node in ast.iter_child_nodes(tree):
        if isinstance(node, ast.FunctionDef) or isinstance(node, ast.AsyncFunctionDef):
            func_doc = extract_docstring(node)
            try:
                func_source = ast.get_source_segment(source, node) or ""
                features = extract_features_from_code(func_source, node.name)
            except:
                features = []

            items.append(CodeItem(
                file_path=relative_path,
                name=node.name,
                item_type='function',
                description=func_doc or 'Module-level function',
                features=features,
                line_number=node.lineno,
            ))

    # If file has no classes/functions, add file-level entry
    if not items:
        items.append(CodeItem(
            file_path=relative_path,
            name=file_path.stem,
            item_type='module',
            description='Module initialization/configuration',
        ))

    return items


def categorize_items(items: List[CodeItem]) -> Dict[str, List[CodeItem]]:
    """Categorize items by functional area based on file path."""
    categories = {
        'Summary Dashboard': [],
        'Jobs & Layout Management': [],
        'Analysis & Detection': [],
        'Track Processing & Extraction': [],
        'Audio Correction': [],
        'Subtitle Processing': [],
        'Orchestrator & Pipeline': [],
        'Post-Processing & Auditors': [],
        'Main Window & Workers': [],
        'Dialogs': [],
        'Widgets & Components': [],
        'Config, Models & Utilities': [],
    }

    for item in items:
        path = item.file_path.lower()

        if 'job_layouts' in path or 'job_discovery' in path:
            categories['Jobs & Layout Management'].append(item)
        elif 'analysis' in path or 'drift' in path or 'sync_stability' in path:
            categories['Analysis & Detection'].append(item)
        elif 'extraction' in path or ('tracks' in path and 'track_widget' not in path):
            categories['Track Processing & Extraction'].append(item)
        elif 'correction' in path and 'auditor' not in path:
            categories['Audio Correction'].append(item)
        elif 'subtitle' in path or 'ocr' in path or 'ass_' in path or 'srt_' in path:
            categories['Subtitle Processing'].append(item)
        elif 'orchestrator' in path or 'pipeline' in path:
            categories['Orchestrator & Pipeline'].append(item)
        elif 'postprocess' in path or 'auditor' in path or 'audit' in path:
            categories['Post-Processing & Auditors'].append(item)
        elif 'main_window' in path or 'worker' in path:
            categories['Main Window & Workers'].append(item)
        elif '_dialog' in path:
            categories['Dialogs'].append(item)
        elif 'widget' in path or 'vsg_qt' in path:
            categories['Widgets & Components'].append(item)
        else:
            categories['Config, Models & Utilities'].append(item)

    return categories


def create_excel_tracker(categories: Dict[str, List[CodeItem]], output_path: Path):
    """Create the Excel tracker workbook."""
    wb = Workbook()

    # Define styles
    header_font = Font(bold=True, size=12, color='FFFFFF')
    header_fill = PatternFill(start_color='2B579A', end_color='2B579A', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Status colors for conditional formatting
    status_colors = {
        'Not Started': 'FFE0E0',  # Light red
        'In Progress': 'FFF3CD',  # Light yellow
        'Completed': 'D4EDDA',    # Light green
        'Testing': 'CCE5FF',      # Light blue
        'Verified': 'C3E6CB',     # Darker green
        'Blocked': 'F5C6CB',      # Pink
    }

    priority_colors = {
        'Critical': 'DC3545',  # Red
        'High': 'FFC107',      # Yellow
        'Medium': '17A2B8',    # Blue
        'Low': '6C757D',       # Gray
    }

    # Column headers
    headers = [
        ('A', 'Done', 6),
        ('B', 'File Path', 40),
        ('C', 'Function/Class/Feature', 35),
        ('D', 'Type', 10),
        ('E', 'Description', 50),
        ('F', 'Features', 35),
        ('G', 'Status', 12),
        ('H', 'Priority', 10),
        ('I', 'Dependencies', 30),
        ('J', 'Notes', 40),
        ('K', 'New Implementation Path', 40),
    ]

    # Create data validations for dropdowns
    status_dv = DataValidation(
        type='list',
        formula1='"Not Started,In Progress,Completed,Testing,Verified,Blocked"',
        allow_blank=True
    )
    status_dv.error = 'Please select a valid status'
    status_dv.errorTitle = 'Invalid Status'

    priority_dv = DataValidation(
        type='list',
        formula1='"Critical,High,Medium,Low"',
        allow_blank=True
    )
    priority_dv.error = 'Please select a valid priority'
    priority_dv.errorTitle = 'Invalid Priority'

    # Create Summary Dashboard first
    ws_summary = wb.active
    ws_summary.title = 'Summary Dashboard'
    create_summary_sheet(ws_summary, categories, headers, header_font, header_fill, header_alignment, thin_border)

    # Create sheets for each category
    sheet_order = [
        'Jobs & Layout Management',
        'Analysis & Detection',
        'Track Processing & Extraction',
        'Audio Correction',
        'Subtitle Processing',
        'Orchestrator & Pipeline',
        'Post-Processing & Auditors',
        'Main Window & Workers',
        'Dialogs',
        'Widgets & Components',
        'Config, Models & Utilities',
    ]

    for sheet_name in sheet_order:
        items = categories.get(sheet_name, [])
        if not items:
            continue

        # Create worksheet
        ws = wb.create_sheet(title=sheet_name[:31])  # Excel sheet name limit

        # Add headers
        for col_letter, header_text, width in headers:
            cell = ws[f'{col_letter}1']
            cell.value = header_text
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            ws.column_dimensions[col_letter].width = width

        # Freeze the header row
        ws.freeze_panes = 'A2'

        # Add data validations
        ws.add_data_validation(status_dv)
        ws.add_data_validation(priority_dv)

        # Add items
        row = 2
        current_file = ""

        for item in sorted(items, key=lambda x: (x.file_path, x.parent_class, x.name)):
            # Add file separator row for better organization
            if item.file_path != current_file:
                current_file = item.file_path

            # Checkbox column (using empty for now - user can add checkmarks)
            ws[f'A{row}'] = ''
            ws[f'A{row}'].alignment = Alignment(horizontal='center')

            # File path
            ws[f'B{row}'] = item.file_path

            # Name with parent class prefix for methods
            name = item.name
            if item.parent_class:
                name = f"{item.parent_class}.{item.name}"
            ws[f'C{row}'] = name

            # Type
            ws[f'D{row}'] = item.item_type.capitalize()

            # Description
            ws[f'E{row}'] = item.description[:200] if item.description else ''

            # Features
            ws[f'F{row}'] = ', '.join(item.features) if item.features else ''

            # Status (default: Not Started)
            ws[f'G{row}'] = 'Not Started'
            status_dv.add(ws[f'G{row}'])

            # Priority (default: Medium)
            priority = 'Medium'
            if item.item_type == 'class':
                priority = 'High'
            elif '__init__' in item.name or 'run' in item.name.lower():
                priority = 'High'
            ws[f'H{row}'] = priority
            priority_dv.add(ws[f'H{row}'])

            # Dependencies
            ws[f'I{row}'] = ''

            # Notes
            ws[f'J{row}'] = ''

            # New Implementation Path
            ws[f'K{row}'] = ''

            # Apply borders
            for col_letter, _, _ in headers:
                ws[f'{col_letter}{row}'].border = thin_border
                ws[f'{col_letter}{row}'].alignment = Alignment(vertical='center', wrap_text=True)

            row += 1

        # Add conditional formatting for status column
        green_fill = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')
        yellow_fill = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')
        red_fill = PatternFill(start_color='FFE0E0', end_color='FFE0E0', fill_type='solid')
        blue_fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')

        # Status conditional formatting rules
        ws.conditional_formatting.add(
            f'G2:G{row}',
            FormulaRule(formula=['$G2="Completed"'], fill=green_fill)
        )
        ws.conditional_formatting.add(
            f'G2:G{row}',
            FormulaRule(formula=['$G2="Verified"'], fill=green_fill)
        )
        ws.conditional_formatting.add(
            f'G2:G{row}',
            FormulaRule(formula=['$G2="In Progress"'], fill=yellow_fill)
        )
        ws.conditional_formatting.add(
            f'G2:G{row}',
            FormulaRule(formula=['$G2="Not Started"'], fill=red_fill)
        )
        ws.conditional_formatting.add(
            f'G2:G{row}',
            FormulaRule(formula=['$G2="Testing"'], fill=blue_fill)
        )
        ws.conditional_formatting.add(
            f'G2:G{row}',
            FormulaRule(formula=['$G2="Blocked"'], fill=PatternFill(start_color='F5C6CB', end_color='F5C6CB', fill_type='solid'))
        )

        # Priority conditional formatting
        ws.conditional_formatting.add(
            f'H2:H{row}',
            FormulaRule(formula=['$H2="Critical"'], fill=PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid'))
        )
        ws.conditional_formatting.add(
            f'H2:H{row}',
            FormulaRule(formula=['$H2="High"'], fill=PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid'))
        )

    # Save workbook
    wb.save(output_path)
    print(f"Tracker saved to: {output_path}")


def create_summary_sheet(ws, categories, headers, header_font, header_fill, header_alignment, thin_border):
    """Create the summary dashboard sheet."""
    # Title
    ws['A1'] = 'VSG Rewrite Tracker - Summary Dashboard'
    ws['A1'].font = Font(bold=True, size=18, color='2B579A')
    ws.merge_cells('A1:F1')

    # Statistics section
    ws['A3'] = 'Overall Progress'
    ws['A3'].font = Font(bold=True, size=14)

    # Category breakdown
    row = 5
    ws['A4'] = 'Category'
    ws['B4'] = 'Total Items'
    ws['C4'] = 'Not Started'
    ws['D4'] = 'In Progress'
    ws['E4'] = 'Completed'
    ws['F4'] = 'Progress %'

    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws[f'{col}4'].font = Font(bold=True)
        ws[f'{col}4'].fill = PatternFill(start_color='E9ECEF', end_color='E9ECEF', fill_type='solid')
        ws[f'{col}4'].border = thin_border

    sheet_names = [
        'Jobs & Layout Management',
        'Analysis & Detection',
        'Track Processing & Extraction',
        'Audio Correction',
        'Subtitle Processing',
        'Orchestrator & Pipeline',
        'Post-Processing & Auditors',
        'Main Window & Workers',
        'Dialogs',
        'Widgets & Components',
        'Config, Models & Utilities',
    ]

    total_items = 0
    for sheet_name in sheet_names:
        items = categories.get(sheet_name, [])
        count = len(items)
        total_items += count

        ws[f'A{row}'] = sheet_name
        ws[f'B{row}'] = count

        # These will be formulas referencing the actual sheets
        safe_sheet_name = sheet_name[:31].replace("'", "''")

        # COUNTIF formulas for status counts
        ws[f'C{row}'] = f'=COUNTIF(\'{safe_sheet_name}\'!G:G,"Not Started")'
        ws[f'D{row}'] = f'=COUNTIF(\'{safe_sheet_name}\'!G:G,"In Progress")'
        ws[f'E{row}'] = f'=COUNTIF(\'{safe_sheet_name}\'!G:G,"Completed")+COUNTIF(\'{safe_sheet_name}\'!G:G,"Verified")'

        # Progress percentage formula
        ws[f'F{row}'] = f'=IF(B{row}>0,ROUND((E{row}/B{row})*100,1)&"%","0%")'

        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws[f'{col}{row}'].border = thin_border

        row += 1

    # Total row
    ws[f'A{row}'] = 'TOTAL'
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'] = f'=SUM(B5:B{row-1})'
    ws[f'C{row}'] = f'=SUM(C5:C{row-1})'
    ws[f'D{row}'] = f'=SUM(D5:D{row-1})'
    ws[f'E{row}'] = f'=SUM(E5:E{row-1})'
    ws[f'F{row}'] = f'=IF(B{row}>0,ROUND((E{row}/B{row})*100,1)&"%","0%")'

    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws[f'{col}{row}'].font = Font(bold=True)
        ws[f'{col}{row}'].fill = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')
        ws[f'{col}{row}'].border = thin_border

    # Column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12

    # Add legend
    row += 3
    ws[f'A{row}'] = 'Status Legend:'
    ws[f'A{row}'].font = Font(bold=True, size=12)

    legend_items = [
        ('Not Started', 'FFE0E0', 'Work has not begun'),
        ('In Progress', 'FFF3CD', 'Currently being worked on'),
        ('Completed', 'D4EDDA', 'Implementation complete'),
        ('Testing', 'CCE5FF', 'Under testing/review'),
        ('Verified', 'C3E6CB', 'Tested and verified working'),
        ('Blocked', 'F5C6CB', 'Blocked by dependency'),
    ]

    row += 1
    for status, color, desc in legend_items:
        ws[f'A{row}'] = status
        ws[f'A{row}'].fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        ws[f'B{row}'] = desc
        row += 1

    # Priority legend
    row += 1
    ws[f'A{row}'] = 'Priority Legend:'
    ws[f'A{row}'].font = Font(bold=True, size=12)

    priority_items = [
        ('Critical', 'F8D7DA', 'Core functionality, must implement first'),
        ('High', 'FFF3CD', 'Important feature, prioritize'),
        ('Medium', 'E2E3E5', 'Standard priority'),
        ('Low', 'E2E3E5', 'Nice to have, can defer'),
    ]

    row += 1
    for priority, color, desc in priority_items:
        ws[f'A{row}'] = priority
        ws[f'A{row}'].fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        ws[f'B{row}'] = desc
        row += 1


def main():
    """Main entry point."""
    base_path = Path('/home/user/Video-Sync-GUI/Reference Only original')
    output_path = Path('/home/user/Video-Sync-GUI/VSG_Rewrite_Tracker.xlsx')

    if not base_path.exists():
        print(f"Error: Base path does not exist: {base_path}")
        return

    print("Scanning Python files...")

    # Find all Python files
    python_files = list(base_path.rglob('*.py'))
    print(f"Found {len(python_files)} Python files")

    # Analyze each file
    all_items = []
    for py_file in python_files:
        items = analyze_python_file(py_file, base_path)
        all_items.extend(items)
        print(f"  Analyzed: {py_file.relative_to(base_path)} ({len(items)} items)")

    print(f"\nTotal items extracted: {len(all_items)}")

    # Categorize items
    categories = categorize_items(all_items)

    for cat_name, cat_items in categories.items():
        if cat_items:
            print(f"  {cat_name}: {len(cat_items)} items")

    # Create Excel tracker
    print("\nCreating Excel tracker...")
    create_excel_tracker(categories, output_path)

    print("\nDone!")
    print(f"Output: {output_path}")


if __name__ == '__main__':
    main()
